from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Dict, List, Optional

import openpyxl




def _get_visible_sheet_names(wb: "openpyxl.Workbook") -> List[str]:
    """Devuelve nombres de hojas visibles en el orden del workbook."""
    names: List[str] = []
    for name in wb.sheetnames:
        ws = wb[name]
        # openpyxl usa 'visible', 'hidden' o 'veryHidden'
        if getattr(ws, "sheet_state", "visible") != "visible":
            continue
        names.append(name)
    return names



def _sheet_to_category(sheet_name: str) -> str:
    """Convierte nombres de hoja a una categoría estable para la UI.

    - Mantiene compatibilidad con las 4 categorías base (singular)
    - Para hojas nuevas, devuelve el nombre tal cual.
    """
    name = (sheet_name or "").strip()
    base_map = {
        "Desayunos": "Desayuno",
        "Comidas": "Comida",
        "Cenas": "Cena",
        "Colaciones": "Colación",
    }
    return base_map.get(name, name)



def _to_str_id(value: Any) -> str:
    """
    Convierte IDs que pueden venir como int/float/str a string limpio.
    Ej: 123456 -> "123456"
    """
    if value is None:
        return ""
    if isinstance(value, (int,)):
        return str(value)
    if isinstance(value, float):
        # Evita "123456.0"
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def read_recipes_from_excel(excel_path: str) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    recipes: List[Dict[str, Any]] = []

    for sheet_name in _get_visible_sheet_names(wb):
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        max_row = ws.max_row or 0

        row = 1
        while row <= max_row:
            tag = ws.cell(row, 1).value

            if tag == "Receta":
                title = ws.cell(row, 2).value
                rid = ws.cell(row, 3).value
                cals = ws.cell(row, 4).value
                prep_time = ws.cell(row, 5).value  # Columna E: Tiempo de preparación (min)

                recipe_id = _to_str_id(rid)
                recipe_title = (str(title).strip() if title is not None else "").strip()
                recipe_cals = float(cals) if isinstance(cals, (int, float)) else None

                # Tiempo de preparación
                recipe_time = None
                if prep_time is not None and str(prep_time).strip() != "":
                    if isinstance(prep_time, (int, float)):
                        minutes = int(prep_time) if float(prep_time).is_integer() else float(prep_time)
                        recipe_time = f"{minutes} min"
                    else:
                        s = str(prep_time).strip()
                        recipe_time = s if "min" in s.lower() else f"{s} min"

                # Preparaci n (fila siguiente esperada)
                prep_text = ""
                if row + 1 <= max_row and ws.cell(row + 1, 1).value == "Preparaci n":
                    prep_val = ws.cell(row + 1, 2).value
                    prep_text = (str(prep_val).strip() if prep_val is not None else "").strip()

                # Header de ingredientes (fila +2 esperada)
                ingredients: List[Dict[str, Any]] = []
                ing_row = row + 3  # datos empiezan 1 fila debajo del header
                # pero primero validamos que exista header en row+2
                if row + 2 <= max_row and ws.cell(row + 2, 1).value == "Ingrediente":
                    ing_row = row + 3

                    while ing_row <= max_row:
                        ing_name = ws.cell(ing_row, 1).value
                        if ing_name is None or str(ing_name).strip() == "":
                            break

                        unit = ws.cell(ing_row, 2).value
                        qty = ws.cell(ing_row, 3).value
                        forma = ws.cell(ing_row, 4).value
                        proceso = ws.cell(ing_row, 5).value

                        ingredients.append({
                            "name": str(ing_name).strip(),
                            "unit": (str(unit).strip() if unit is not None else ""),
                            "qty": qty,
                            "forma": (str(forma).strip() if forma is not None else ""),
                            "proceso": (str(proceso).strip() if proceso is not None else ""),
                        })
                        ing_row += 1

                # Construimos un DTO  compatible  con tu UI actual (con defaults)
                # Nota: img lo resolveremos en Paso 1.3, por ahora lo dejamos null.
                recipes.append({
                    "id": recipe_id,
                    "title": recipe_title,
                    "category": _sheet_to_category(sheet_name) if sheet_name.endswith("s") else sheet_name,  # "Desayuno", etc.
                    "time": recipe_time,
                    "cals": recipe_cals,
                    "price": None,
                    "img": None,
                    "ingredients": ingredients,
                    "prep": prep_text
                })

                # Saltamos hasta despu s de la receta (buscamos primera fila vac a tras ingredientes)
                # Si no hubo ingredientes, avanzamos m nimo 1.
                row = max(row + 1, ing_row + 1)
                continue

            row += 1

    return recipes
