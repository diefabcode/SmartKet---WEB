import argparse
import json
import os
import secrets
import math
import re
from datetime import datetime, timezone
import smtplib
from email.message import EmailMessage
from pathlib import Path

from flask import Flask, jsonify, send_file, request, make_response

import openpyxl

from services.excel_reader import read_recipes_from_excel
from services.image_resolver import build_image_index


# Env overrides (preferidas sobre config.json)
ENV_EXCEL_PATH = "SMARTKET_EXCEL_PATH"
ENV_RESOURCES_DIR = "SMARTKET_RESOURCES_DIR"


# ----------------------------
# Ingredientes.xlsx helpers (cotización de pedido)
# ----------------------------

_CANON_UNIT_MAP = {
    "g": "gramos",
    "gramo": "gramos",
    "gramos": "gramos",
    "kg": "kilogramo",
    "kilogramo": "kilogramo",
    "kilogramos": "kilogramo",
    "ml": "mililitro",
    "mililitro": "mililitro",
    "mililitros": "mililitro",
    "l": "litro",
    "litro": "litro",
    "litros": "litro",
    "cucharada": "cucharada",
    "cucharadas": "cucharada",
    "cda": "cucharada",
    "cdas": "cucharada",
    "cucharadita": "cucharadita",
    "cucharaditas": "cucharadita",
    "cdta": "cucharadita",
    "cdita": "cucharadita",
    "taza": "taza",
    "tazas": "taza",
}

# Conversión simple (fase 1).
# Se asume para líquidos: cucharada=15ml, cucharadita=5ml, taza=240ml.
_UNIT_CONVERSIONS = {
    ("kilogramo", "gramos"): 1000.0,
    ("gramos", "kilogramo"): 1.0 / 1000.0,
    ("litro", "mililitro"): 1000.0,
    ("mililitro", "litro"): 1.0 / 1000.0,
    ("cucharada", "mililitro"): 15.0,
    ("cucharadita", "mililitro"): 5.0,
    ("taza", "mililitro"): 240.0,
}


def _norm_unit(u):
    u = (u or "").strip().lower()
    u = u.replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")
    return _CANON_UNIT_MAP.get(u, u)


def _try_convert_qty(qty, from_unit, to_unit):
    fu = _norm_unit(from_unit)
    tu = _norm_unit(to_unit)
    if fu == tu:
        return float(qty)
    key = (fu, tu)
    if key in _UNIT_CONVERSIONS:
        return float(qty) * float(_UNIT_CONVERSIONS[key])
    return None


def _split_pipe(cell_val):
    if cell_val is None:
        return []
    s = str(cell_val).strip()
    if not s:
        return []
    return [x.strip() for x in s.split("|")]


def _to_float_safe(v):
    if v is None:
        return None
    try:
        s = str(v).strip()
        if not s:
            return None
        s = s.replace(",", ".")
        return float(s)
    except Exception:
        return None


def _derive_ingredientes_xlsx_path(alimentos_excel_path):
    """
    Si el config solo trae Alimentos.xlsx, inferimos Ingredientes.xlsx en el mismo folder.
    """
    p = Path(alimentos_excel_path)
    candidate = p.with_name("Ingredientes.xlsx")
    return str(candidate)


def _load_ingredients_catalog(ingredientes_xlsx_path):
    """
    Catálogo estricto (pipe por índice):

    Regla NO negociable:
      Si hay N marcas, debe haber N valores para:
        - PRECIO DE COMPRA
        - PRECIO DE VENTA
        - PRESENTACION
        - UNIDAD
        - CANTIDAD
        - GRANEL (SI/NO)
        - MULTIPLOS (o "na")
        - LUGAR

    NOTA:
      CATEGORÍA / FORMA / PROCESO pueden traer "|" pero NO tienen correspondencia con N marcas.
      (Se manejan en meta/otros flujos, no aquí.)
    """
    issues = []
    if not ingredientes_xlsx_path or not os.path.isfile(ingredientes_xlsx_path):
        return {}, [f"No existe Ingredientes.xlsx en: {ingredientes_xlsx_path}"]

    wb = openpyxl.load_workbook(ingredientes_xlsx_path, data_only=True)
    ws = wb.active

    header = {str(ws.cell(1, c).value).strip(): c for c in range(1, 60) if ws.cell(1, c).value}
    required = [
        "INGREDIENTE",
        "MARCA",
        "PRECIO DE COMPRA",
        "PRECIO DE VENTA",
        "PRESENTACION",
        "UNIDAD",
        "CANTIDAD",
        "GRANEL",
        "MULTIPLOS",
        "LUGAR",
    ]
    for k in required:
        if k not in header:
            issues.append(f"Falta columna requerida en Ingredientes.xlsx: {k}")

    if issues:
        return {}, issues

    col_ing = header["INGREDIENTE"]
    col_marca = header["MARCA"]
    col_buy = header["PRECIO DE COMPRA"]
    col_sale = header["PRECIO DE VENTA"]
    col_pres = header["PRESENTACION"]
    col_unit = header["UNIDAD"]
    col_qty = header["CANTIDAD"]
    col_granel = header["GRANEL"]
    col_mult = header["MULTIPLOS"]
    col_lugar = header["LUGAR"]

    catalog = {}
    max_row = ws.max_row or 1

    for r in range(2, max_row + 1):
        ing = ws.cell(r, col_ing).value
        if ing is None or str(ing).strip() == "":
            continue
        name = str(ing).strip()

        marcas = _split_pipe(ws.cell(r, col_marca).value)
        buys = _split_pipe(ws.cell(r, col_buy).value)
        sales = _split_pipe(ws.cell(r, col_sale).value)
        pres = _split_pipe(ws.cell(r, col_pres).value)
        units = _split_pipe(ws.cell(r, col_unit).value)
        qtys = _split_pipe(ws.cell(r, col_qty).value)
        granels = _split_pipe(ws.cell(r, col_granel).value)
        multiples = _split_pipe(ws.cell(r, col_mult).value)
        lugares = _split_pipe(ws.cell(r, col_lugar).value)

        # Validación estricta por correspondencia (pipe por índice)
        arrays = {
            "MARCA": marcas,
            "PRECIO DE COMPRA": buys,
            "PRECIO DE VENTA": sales,
            "PRESENTACION": pres,
            "UNIDAD": units,
            "CANTIDAD": qtys,
            "GRANEL": granels,
            "MULTIPLOS": multiples,
            "LUGAR": lugares,
        }

        lengths = {k: len(v) for k, v in arrays.items()}
        n = lengths.get("MARCA", 0)

        if n == 0:
            issues.append(f"[{name}] Sin MARCA (no hay ofertas).")
            continue

        bad = [k for k, ln in lengths.items() if ln != n]
        if bad:
            issues.append(
                f"[{name}] Ofertas desalineadas (pipe por índice). "
                f"Se esperaba N={n} en todas las columnas. Diferentes: {', '.join([f'{k}={lengths[k]}' for k in bad])}."
            )
            # Estricto: no se construye este ingrediente.
            continue

        offers = []
        for i in range(n):
            bulk_s = (granels[i] or "").strip().upper()
            is_bulk_offer = (bulk_s == "SI")

            mult_val = _to_float_safe(multiples[i])
            # "na" o vacío -> None
            if mult_val is not None and float(mult_val) <= 0:
                mult_val = None

            place_val = (lugares[i] or "").strip()
            if not place_val:
                issues.append(f"[{name}] LUGAR vacío en índice {i+1}.")
                continue

            offer = {
                "index": i,
                "brand": (marcas[i] or "").strip(),
                "buy": _to_float_safe(buys[i]),
                "sale": _to_float_safe(sales[i]),
                "presentation": (pres[i] or "").strip(),
                "unit": (units[i] or "").strip(),
                "qty": _to_float_safe(qtys[i]),
                "bulk": is_bulk_offer,
                "multiple": mult_val,
                "place": place_val,
            }

            if not offer["brand"]:
                issues.append(f"[{name}] MARCA vacía en índice {i+1}.")
                continue
            if offer["sale"] is None or offer["buy"] is None:
                issues.append(f"[{name}] Precio inválido en índice {i+1}. Revisa COMPRA/VENTA.")
                continue
            if offer["qty"] is None or offer["qty"] <= 0:
                issues.append(f"[{name}] CANTIDAD inválida en índice {i+1}.")
                continue
            if not offer["unit"]:
                issues.append(f"[{name}] UNIDAD vacía en índice {i+1}.")
                continue

            offers.append(offer)

        if not offers:
            issues.append(f"[{name}] Sin ofertas válidas (revisa filas/valores).")
            continue

        catalog[name] = {"offers": offers}

    return catalog, issues

def _load_ingredients_meta(ingredientes_xlsx_path):
    """
    Meta ligera para UI.

    Devuelve:
      meta[name] = { "category": "<CATEGORÍA>" }
    """
    issues = []
    if not ingredientes_xlsx_path or not os.path.isfile(ingredientes_xlsx_path):
        return {}, [f"No existe Ingredientes.xlsx en: {ingredientes_xlsx_path}"]

    wb = openpyxl.load_workbook(ingredientes_xlsx_path, data_only=True)
    ws = wb.active

    # Leer encabezados (fila 1)
    header = {}
    max_col = ws.max_column or 1
    for c in range(1, max_col + 1):
        v = ws.cell(1, c).value
        if v is None:
            continue
        key = str(v).strip().upper()
        if key:
            header[key] = c

    if "INGREDIENTE" not in header:
        return {}, ["Falta columna requerida en Ingredientes.xlsx: INGREDIENTE"]

    col_ing = header["INGREDIENTE"]
    col_cat = header.get("CATEGORÍA") or header.get("CATEGORIA")  # tolerante a acento

    if not col_cat:
        # No romper: simplemente devolvemos meta vacío.
        return {}, ["Falta columna CATEGORÍA en Ingredientes.xlsx (no se puede categorizar)."]

    meta = {}
    max_row = ws.max_row or 1
    for r in range(2, max_row + 1):
        ing = ws.cell(r, col_ing).value
        if ing is None or str(ing).strip() == "":
            continue
        name = str(ing).strip()

        cat = ws.cell(r, col_cat).value
        cat_s = str(cat).strip() if cat is not None else ""
        meta[name] = {"category": cat_s}

    return meta, issues


def _aggregate_plan_ingredients(payload):
    """
    Devuelve dict ingredient_name -> list de usos (qty, unit, portions, recipe_id, recipe_title)
    """
    plan = payload.get("plan") or {}
    agg = {}

    if isinstance(plan, list):
        day_items_iter = enumerate(plan)
    elif isinstance(plan, dict):
        def _k(x):
            try:
                return int(x)
            except Exception:
                return 10**9
        day_items_iter = ((k, plan[k]) for k in sorted(plan.keys(), key=_k))
    else:
        return agg

    for _, items in day_items_iter:
        if not items:
            continue
        for it in items:
            try:
                portions = int(it.get("portions", 1) or 1)
            except Exception:
                portions = 1

            ingredients = it.get("ingredients") or []
            for ing in ingredients:
                name = (ing.get("name") or "").strip()
                if not name:
                    continue
                unit = (ing.get("unit") or "").strip()
                qty = ing.get("qty", None)
                try:
                    qty_f = float(qty)
                except Exception:
                    continue

                use = {
                    "qty": qty_f,
                    "unit": unit,
                    "portions": portions,
                    "recipe_id": str(it.get("id", "")).strip(),
                    "recipe_title": (it.get("title") or "").strip(),
                }
                agg.setdefault(name, []).append(use)

    return agg


def _norm_ing_name(s: str) -> str:
    return (str(s or "").strip().lower())

def _quote_sellable_items(agg, catalog, offer_overrides=None):
    """
    Convierte 'usos de receta' en 'items vendibles' y total (precio de venta).

    Soporta overrides por ingrediente:
      offer_overrides = { "Chorizo": 0, "Leche": 1, ... }
    donde el valor es el índice dentro de catalog[name]["offers"].

    NOTA:
      Si no se envían overrides, el comportamiento base se mantiene: elige la oferta más conveniente.
    """
    items = []
    issues = []
    total = 0.0

    # Mapa normalizado para tolerar mayúsculas/espacios
    overrides_norm = {}
    if isinstance(offer_overrides, dict):
        for k, v in offer_overrides.items():
            try:
                overrides_norm[_norm_ing_name(k)] = int(v)
            except Exception:
                continue

    # Índice normalizado del catálogo (para buscar por nombre tolerante)
    catalog_norm = {_norm_ing_name(k): k for k in catalog.keys()}

    for ing_name, uses in agg.items():
        key_exact = ing_name if ing_name in catalog else None
        key_norm = catalog_norm.get(_norm_ing_name(ing_name))
        key = key_exact or key_norm

        if not key or key not in catalog:
            issues.append(f"Ingrediente no encontrado en catálogo: {ing_name}")
            continue

        entry = catalog[key]
        offers = entry.get("offers") or []
        if not offers:
            issues.append(f"Sin ofertas válidas para ingrediente: {ing_name}")
            continue

        def required_in_unit(target_unit):
            total_req = 0.0
            for u in uses:
                q = float(u["qty"]) * float(u.get("portions", 1))
                conv = _try_convert_qty(q, u.get("unit", ""), target_unit)
                if conv is None:
                    return None
                total_req += conv
            return total_req

        evals = []
        for idx, off in enumerate(offers):
            off_unit = off.get("unit")
            off_qty = off.get("qty")
            off_sale = off.get("sale")
            if not off_unit or off_qty is None or off_sale is None:
                continue

            req = required_in_unit(off_unit)
            if req is None:
                continue

            canon_unit = _norm_unit(off_unit)
            canon_qty = float(off_qty)
            if canon_unit == "litro":
                canon_qty = float(off_qty) * 1000.0
                canon_unit = "mililitro"
            elif canon_unit == "kilogramo":
                canon_qty = float(off_qty) * 1000.0
                canon_unit = "gramos"

            sale_per_canon_unit = float(off_sale) / canon_qty if canon_qty > 0 else float("inf")

            # Requerimiento en cantidad canónica
            req_canon_qty = float(req)
            u_norm = _norm_unit(off_unit)
            if u_norm == "litro":
                req_canon_qty = float(req) * 1000.0
            elif u_norm == "kilogramo":
                req_canon_qty = float(req) * 1000.0

            # Cálculo de compra/venta según modo
            is_bulk_offer = bool(off.get("bulk", False))
            multiple = off.get("multiple", None)

            if is_bulk_offer:
                sold_canon_qty = req_canon_qty
                if isinstance(multiple, (int, float)) and float(multiple) > 0:
                    mval = float(multiple)
                    sold_canon_qty = float(math.ceil(req_canon_qty / mval) * mval)
                line_total = sold_canon_qty * sale_per_canon_unit
                waste = sold_canon_qty - req_canon_qty
                packs = None
            else:
                pkg_canon_qty = canon_qty
                packs = int(math.ceil(req_canon_qty / pkg_canon_qty)) if pkg_canon_qty > 0 else 0
                line_total = float(off_sale) * packs
                waste = (packs * pkg_canon_qty) - req_canon_qty

            evals.append({
                "offer_index": int(off.get("index", idx)),
                "offer": off,
                "req_in_offer_unit": float(req),
                "canon_unit": canon_unit,
                "canon_pkg_qty": canon_qty,
                "sale_per_canon_unit": sale_per_canon_unit,
                "req_canon_qty": req_canon_qty,
                "sold_canon_qty": (sold_canon_qty if is_bulk_offer else (packs * canon_qty)),
                "packages_needed": packs,
                "line_total": float(line_total),
                "waste_qty": float(waste),
                "sell_mode": "bulk" if is_bulk_offer else "package",
            })

        if not evals:
            issues.append(f"No se pudo convertir unidades para ingrediente: {ing_name} (unidad receta vs catálogo)")
            continue

        # Default: elegir la opción más conveniente por costo total (y menos desperdicio como desempate)
        chosen_default = min(evals, key=lambda e: (e["line_total"], e["waste_qty"], e["sale_per_canon_unit"]))
        default_idx = chosen_default["offer_index"]

        # Override (si existe y es válido)
        chosen = chosen_default
        ovr = overrides_norm.get(_norm_ing_name(ing_name))
        override_applied = False
        if isinstance(ovr, int) and 0 <= ovr < len(offers):
            # Buscar eval con ese offer_index (puede coincidir con i)
            cand = next((e for e in evals if e["offer_index"] == ovr), None)
            if cand is not None:
                chosen = cand
                override_applied = True
            else:
                issues.append(f"[{ing_name}] Override inválido por unidad no convertible para índice {ovr}. Se usó default.")

        off = chosen["offer"]

        # Resumen de opciones para UI (sin cálculos)
        offers_summary = [{
            "index": int(o.get("index", i)),
            "brand": o.get("brand"),
            "buy": o.get("buy"),
            "sale": o.get("sale"),
            "presentation": o.get("presentation"),
            "unit": o.get("unit"),
            "qty": o.get("qty"),
            "bulk": bool(o.get("bulk", False)),
            "multiple": o.get("multiple"),
            "place": o.get("place"),
        } for i, o in enumerate(offers)]

        line_total = chosen["line_total"]
        total += line_total

        # Item para UI
        item = {
            "ingredient": ing_name,
            "sell_mode": chosen["sell_mode"],
            "selected_offer_index": int(chosen["offer_index"]),
            "default_offer_index": int(default_idx),
            "override_applied": bool(override_applied),
            "offer_brand": off.get("brand"),
            "offer_presentation": off.get("presentation"),
            "offer_place": off.get("place"),
            "offers": offers_summary,
            "required_qty": round(chosen["req_canon_qty"], 6),
            "waste_qty": round(chosen["waste_qty"], 6),
            "line_total": round(line_total, 2),
        }

        if chosen["sell_mode"] == "bulk":
            item.update({
                "bulk_unit": chosen["canon_unit"],
                "sold_qty": round(chosen["sold_canon_qty"], 6),
                "unit_price": round(chosen["sale_per_canon_unit"], 6),
                "multiple": off.get("multiple"),
            })
        else:
            item.update({
                "package_unit": chosen["canon_unit"],
                "package_qty": round(chosen["canon_pkg_qty"], 6),
                "packages_needed": int(chosen["packages_needed"] or 0),
                "unit_price": round(float(off.get("sale") or 0.0), 2),
            })

        items.append(item)

    items.sort(key=lambda x: (x.get("sell_mode", ""), x.get("ingredient", "").lower()))
    return items, round(total, 2), issues

def _iso_date_from_dynamic_day(full_date: str) -> str:
    """Convierte dynamicDays[i].fullDate a 'YYYY-MM-DD'. Acepta ISO con 'Z'."""
    if not full_date:
        raise ValueError("fullDate vacío")
    s = str(full_date).strip()
    if s.endswith("Z"):
        s = s[:-1] + "+00:00"
    # fromisoformat requiere offset explícito si hay zona
    dtv = datetime.fromisoformat(s)
    return dtv.date().isoformat()


def _safe_days_from_dynamic_days(dynamic_days) -> list:
    """Devuelve lista ['YYYY-MM-DD', ...] con el mismo orden que dynamicDays."""
    if not isinstance(dynamic_days, list):
        return []
    out = []
    for d in dynamic_days:
        try:
            out.append(_iso_date_from_dynamic_day(d.get("fullDate")))
        except Exception:
            # fallback: intenta parsear dateStr si ya viniera como YYYY-MM-DD
            fd = (d.get("fullDate") or d.get("dateStr") or "").strip()
            if re.match(r"^\d{4}-\d{2}-\d{2}$", fd):
                out.append(fd)
            else:
                out.append("")
    return out


def _norm_key(s: str) -> str:
    return str(s or "").strip().lower()

# ----------------------------
# PurchaseSnapshot helpers (contrato fuerte para WPF)
# ----------------------------

def _strip_accents(s: str) -> str:
    s = str(s or "")
    # Mantener simple (sin dependencias externas)
    return (s
            .replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")
            .replace("Á", "a").replace("É", "e").replace("Í", "i").replace("Ó", "o").replace("Ú", "u")
            .replace("ñ", "n").replace("Ñ", "n"))


def _collapse_spaces(s: str) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip())


def _norm_key_part(s: str) -> str:
    """Normaliza para keys: sin acentos, sin dobles espacios, lower-case."""
    s = _strip_accents(str(s or ""))
    s = _collapse_spaces(s)
    return s.lower()


def _num_to_key(v) -> str:
    """Convierte números a string estable para key (sin .0)."""
    if v is None:
        return "na"
    try:
        f = float(v)
        if f.is_integer():
            return str(int(f))
        # evita notación científica si se puede
        return ("%s" % f).rstrip("0").rstrip(".")
    except Exception:
        return _norm_key_part(str(v))

def _maybe_int_num(v):
    """Devuelve int si es entero, si no float; None si None."""
    if v is None:
        return None
    try:
        f = float(v)
        if f.is_integer():
            return int(f)
        return float(f)
    except Exception:
        return v


def _human_num_str(v) -> str:
    """Formato humano estable (sin .0)."""
    if v is None:
        return ""
    try:
        f = float(v)
        if f.is_integer():
            return str(int(f))
        return ("%s" % f).rstrip("0").rstrip(".")
    except Exception:
        return str(v)


def _build_display_label_package(ingredient: str, brand: str, presentation: str, pack_qty, unit: str) -> str:
    pq = _human_num_str(_maybe_int_num(pack_qty))
    return f"{ingredient} | {brand} | {presentation} | {pq} {unit}".strip()


def _build_display_label_bulk(ingredient: str, brand: str, presentation: str, unit: str, multiple_val) -> str:
    base = f"{ingredient} | {brand} | {presentation} | {unit}".strip()
    if multiple_val is not None:
        mv = _human_num_str(_maybe_int_num(multiple_val))
        base = f"{base} | múltiplo {mv}"
    return base


def _build_quantity_label_package(packages_needed: int, buy_qty, unit: str) -> str:
    base = f"Cantidad: {int(packages_needed)} paquetes"
    if buy_qty is not None and unit:
        bq = _human_num_str(_maybe_int_num(buy_qty))
        base = f"{base} ({bq} {unit} total)"
    return base


def _build_quantity_label_bulk(buy_qty, unit: str) -> str:
    bq = _human_num_str(_maybe_int_num(buy_qty))
    return f"Cantidad: {bq} {unit}".strip()


def _unit_for_snapshot(canon_unit: str) -> str:
    """Snapshot usa unidades humanas (ej: ml, gramos)."""
    u = (canon_unit or "").strip().lower()
    if u == "mililitro":
        return "ml"
    if u == "litro":
        return "l"
    if u == "kilogramo":
        return "kg"
    return u


def _canon_offer_unit_qty(offer_unit, offer_qty):
    """Canoniza qty de la oferta (litro->ml, kilogramo->gramos) igual que la cotización."""
    ou = offer_unit
    oq = offer_qty
    try:
        oqf = float(oq)
    except Exception:
        return None, None

    canon_unit = _norm_unit(ou)
    canon_qty = float(oqf)

    if canon_unit == "litro":
        canon_unit = "mililitro"
        canon_qty = float(oqf) * 1000.0
    elif canon_unit == "kilogramo":
        canon_unit = "gramos"
        canon_qty = float(oqf) * 1000.0

    return canon_unit, canon_qty


def _build_purchase_line_key_package(place, ingredient, brand, presentation, unit, pack_qty) -> str:
    return "|".join([
        _norm_key_part(place),
        _norm_key_part(ingredient),
        _norm_key_part(brand),
        _norm_key_part(presentation),
        _norm_key_part(unit),
        _num_to_key(pack_qty),
        "package",
    ])


def _build_purchase_line_key_bulk(place, ingredient, brand, presentation, unit, multiple) -> str:
    mult = _num_to_key(multiple)
    return "|".join([
        _norm_key_part(place),
        _norm_key_part(ingredient),
        _norm_key_part(brand),
        _norm_key_part(presentation),
        _norm_key_part(unit),
        f"multiple={mult}",
        "bulk",
    ])


def _find_selected_offer(item: dict):
    sel = item.get("selected_offer_index")
    offers = item.get("offers") or []
    for o in offers:
        try:
            if int(o.get("index")) == int(sel):
                return o
        except Exception:
            continue
    return None


def _build_purchase_snapshot_from_quote_items(items: list) -> dict:
    """Construye purchaseSnapshot.lines[] a partir de quoteSnapshot.items[].

    Estrategia de errores: omite líneas sin oferta válida (ya quedan registradas en quoteSnapshot.issues).
    """
    now_utc = datetime.now(timezone.utc).isoformat()
    out = {
        "schema_version": 2,
        "generated_at_utc": now_utc,
        "semantics": {
            "package_mode": {
                "primary_quantity_field": "packages_needed",
                "buy_qty_definition": "packages_needed * offer.pack_qty"
            },
            "bulk_mode": {
                "primary_quantity_field": "buy_qty",
                "buy_qty_definition": "required_qty rounded up to offer.multiple when present"
            }
        },
        "lines": [],
    }

    if not isinstance(items, list):
        return out

    for it in items:
        if not isinstance(it, dict):
            continue

        ingredient = (it.get("ingredient") or "").strip()
        sell_mode = (it.get("sell_mode") or "").strip().lower()
        selected_offer_index = it.get("selected_offer_index")

        off = _find_selected_offer(it)
        if not off:
            # Sin oferta -> omitimos (issues ya lo reporta en quoteSnapshot)
            continue

        place = (it.get("offer_place") or off.get("place") or "").strip()
        brand = (it.get("offer_brand") or off.get("brand") or "").strip()
        presentation = (it.get("offer_presentation") or off.get("presentation") or "").strip()

        # Canonizar unit/pack_qty para que el snapshot sea estable
        canon_unit, canon_qty = _canon_offer_unit_qty(off.get("unit"), off.get("qty"))
        if canon_unit is None or canon_qty is None:
            continue

        snap_unit = _unit_for_snapshot(canon_unit)

        required_qty = it.get("required_qty")
        try:
            required_qty_f = float(required_qty)
        except Exception:
            continue

        multiple = off.get("multiple")
        multiple_val = None
        try:
            if multiple is not None and str(multiple).strip() != "":
                mv = float(multiple)
                if mv > 0:
                    multiple_val = mv
        except Exception:
            multiple_val = None

        if sell_mode == "package":
            pack_qty = canon_qty
            try:
                packages_needed = int(it.get("packages_needed") or 0)
            except Exception:
                packages_needed = 0
            if packages_needed <= 0:
                # Recalcular por seguridad
                packages_needed = int(math.ceil(required_qty_f / float(pack_qty))) if float(pack_qty) > 0 else 0

            buy_qty = float(packages_needed) * float(pack_qty)

            purchase_line_key = _build_purchase_line_key_package(
                place=place,
                ingredient=ingredient,
                brand=brand,
                presentation=presentation,
                unit=snap_unit,
                pack_qty=pack_qty,
            )

            line = {
                "purchase_line_key": purchase_line_key,
                "ingredient": ingredient,
                "place": place,
                "sell_mode": "package",
                "selected_offer_index": int(selected_offer_index) if selected_offer_index is not None else None,
                "offer": {
                    "brand": brand,
                    "presentation": presentation,
                    "unit": snap_unit,
                    "pack_qty": _maybe_int_num(pack_qty),
                    "bulk": False,
                    "multiple": None,
                },
                "required_qty": float(required_qty_f),
                "packages_needed": int(packages_needed),
                "buy_qty": float(buy_qty),
                "rounding_rule": "ceil(required_qty / pack_qty)",
                "display_label": _build_display_label_package(ingredient, brand, presentation, pack_qty, snap_unit),
                "quantity_label": _build_quantity_label_package(int(packages_needed), buy_qty, snap_unit),
            }
            out["lines"].append(line)

        elif sell_mode == "bulk":
            # buy_qty: usar sold_qty ya calculado (redondeo por múltiplo si aplica)
            sold_qty = it.get("sold_qty")
            try:
                buy_qty = float(sold_qty) if sold_qty is not None else float(required_qty_f)
            except Exception:
                buy_qty = float(required_qty_f)

            mult_for_key = multiple_val if multiple_val is not None else "na"

            purchase_line_key = _build_purchase_line_key_bulk(
                place=place,
                ingredient=ingredient,
                brand=brand,
                presentation=presentation,
                unit=snap_unit,
                multiple=mult_for_key,
            )

            rounding_rule = "no_rounding"
            if multiple_val is not None:
                rounding_rule = "round_up_to_multiple(required_qty, multiple)"

            line = {
                "purchase_line_key": purchase_line_key,
                "ingredient": ingredient,
                "place": place,
                "sell_mode": "bulk",
                "selected_offer_index": int(selected_offer_index) if selected_offer_index is not None else None,
                "offer": {
                    "brand": brand,
                    "presentation": presentation,
                    "unit": snap_unit,
                    # Para bulk mantenemos pack_qty como el tamaño de la oferta (canonizado) por trazabilidad
                    "pack_qty": _maybe_int_num(canon_qty),
                    "bulk": True,
                    "multiple": int(multiple_val) if (isinstance(multiple_val, float) and multiple_val.is_integer()) else multiple_val,
                },
                "required_qty": float(required_qty_f),
                "packages_needed": None,
                "buy_qty": float(buy_qty),
                "rounding_rule": rounding_rule,
                "display_label": _build_display_label_bulk(ingredient, brand, presentation, snap_unit, multiple_val),
                "quantity_label": _build_quantity_label_bulk(buy_qty, snap_unit),
            }
            out["lines"].append(line)

        else:
            # modo desconocido -> omitimos
            continue

    return out


def _build_plan_for_quote(plan, excluded_ingredients) -> list:
    """
    Convierte el plan 'rico' que manda la WEB (con ingredients_raw, assignedMeal, etc.)
    a un plan mínimo que entiende /api/orders/quote:
      plan: [ [ { id,title,portions,ingredients:[{name,unit,qty}]} ] ]
    Además aplica exclusiones (post-exclusión).
    """
    excl = set(_norm_key(x) for x in (excluded_ingredients or []) if str(x or "").strip())
    # Ordenar días
    if isinstance(plan, list):
        day_items_iter = enumerate(plan)
        max_day = len(plan)
    elif isinstance(plan, dict):
        def _k(x):
            try:
                return int(x)
            except Exception:
                return 10**9
        keys = sorted(plan.keys(), key=_k)
        day_items_iter = ((int(k) if str(k).isdigit() else k, plan[k]) for k in keys)
        max_day = len(keys)
    else:
        return []

    # Construimos lista por índice (0..n-1) para estabilidad
    out = [[] for _ in range(max_day)]
    for day_idx, items in day_items_iter:
        try:
            di = int(day_idx)
        except Exception:
            continue
        if di < 0 or di >= len(out):
            continue

        if not isinstance(items, list):
            continue

        for it in items:
            if not isinstance(it, dict):
                continue
            rid = str(it.get("id", "")).strip()
            title = (it.get("title") or "").strip()
            try:
                portions = int(it.get("portions", 1) or 1)
            except Exception:
                portions = 1

            raw = it.get("ingredients_raw") or []
            ingredients = []
            if isinstance(raw, list):
                for ing in raw:
                    if not isinstance(ing, dict):
                        continue
                    name = (ing.get("name") or "").strip()
                    if not name:
                        continue
                    if _norm_key(name) in excl:
                        continue
                    unit = (ing.get("unit") or "").strip()
                    qty = ing.get("qty", None)
                    ingredients.append({
                        "name": name,
                        "unit": unit,
                        "qty": qty,
                    })

            out[di].append({
                "id": rid,
                "title": title,
                "portions": portions,
                "ingredients": ingredients,
            })

    return out


def _build_recipe_blocks(plan, dynamic_days, excluded_ingredients) -> list:
    """
    Construye recipeBlocks[] post-exclusión:
      { day, slot, recipe{...}, portions, prep, ingredients[] escalados }
    """
    days = _safe_days_from_dynamic_days(dynamic_days)
    excl = set(_norm_key(x) for x in (excluded_ingredients or []) if str(x or "").strip())

    # Normaliza plan a dict por índice para acceder con dayIdx
    plan_by_day = {}
    if isinstance(plan, list):
        for i, items in enumerate(plan):
            plan_by_day[i] = items
    elif isinstance(plan, dict):
        for k, v in plan.items():
            try:
                plan_by_day[int(k)] = v
            except Exception:
                continue

    blocks = []
    # Recorremos por len(days) para que el pedido sea estable por fechas reales
    for day_idx, day_iso in enumerate(days):
        items = plan_by_day.get(day_idx) or []
        if not isinstance(items, list):
            continue

        for it in items:
            if not isinstance(it, dict):
                continue

            slot = (it.get("assignedMeal") or "").strip()
            if slot not in ("Desayuno", "Comida", "Cena", "Colación"):
                # Si algo raro entra, lo dejamos como string (pero nunca null)
                slot = slot or "Comida"

            try:
                portions = int(it.get("portions", 1) or 1)
            except Exception:
                portions = 1

            rid = str(it.get("id", "")).strip()
            title = (it.get("title") or "").strip()
            prep = (it.get("prep") or "").strip()

            raw = it.get("ingredients_raw") or []
            ingredients_out = []
            if isinstance(raw, list):
                for ing in raw:
                    if not isinstance(ing, dict):
                        continue
                    name = (ing.get("name") or "").strip()
                    if not name:
                        continue
                    if _norm_key(name) in excl:
                        continue

                    unit = (ing.get("unit") or "").strip()
                    forma = (ing.get("forma") or "").strip()
                    proceso = (ing.get("proceso") or "").strip()

                    qty = ing.get("qty", None)
                    qty_scaled = None
                    if qty is not None and str(qty).strip() != "":
                        try:
                            qty_scaled = float(qty) * float(portions)
                            # Redondeo suave para evitar basura flotante
                            qty_scaled = round(qty_scaled, 6)
                        except Exception:
                            qty_scaled = qty

                    ingredients_out.append({
                        "name": name,
                        "unit": unit,
                        "qty": qty_scaled,
                        "form": forma,
                        "process": proceso,
                    })

            blocks.append({
                "day": day_iso,
                "day_index": day_idx,
                "slot": slot,
                "recipe": {
                    "id": rid,
                    "title": title,
                },
                "portions": portions,
                "prep": prep,
                "ingredients": ingredients_out,
            })

    return blocks


def load_config(path: str) -> dict:
    if not path:
        raise ValueError("Config path vacío.")
    if not os.path.isfile(path):
        raise FileNotFoundError(f"No existe el archivo de config: {path}")

    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Defaults
    data.setdefault("host", "127.0.0.1")
    data.setdefault("port", 5050)

    # Validación mínima de llaves requeridas
    for key in ["excel_path", "resources_dir"]:
        if key not in data or not str(data[key]).strip():
            raise ValueError(f"Falta '{key}' en config o está vacío.")

    return data


def create_app(config: dict) -> Flask:
    app = Flask(__name__)
    app.config["SMARTKET_CONFIG"] = config

    def _get_effective_paths(cfg: dict):
        """
        Resuelve rutas efectivas para Excel/Recursos.
        Prioridad: Environment Variables > config.json
        """
        env_excel = (os.environ.get(ENV_EXCEL_PATH) or "").strip()
        env_resdir = (os.environ.get(ENV_RESOURCES_DIR) or "").strip()

        excel_path = env_excel or str(cfg.get("excel_path", "") or "").strip()
        resources_dir = env_resdir or str(cfg.get("resources_dir", "") or "").strip()

        info = {
            "excel_path": excel_path,
            "resources_dir": resources_dir,
            "excel_source": "env" if env_excel else "config",
            "resources_source": "env" if env_resdir else "config",
        }

        issues = []
        if not excel_path:
            issues.append(f"Ruta de Excel vacía. Define {ENV_EXCEL_PATH} o 'excel_path' en config.")
        elif not os.path.isfile(excel_path):
            issues.append(f"No existe el archivo Excel en: {excel_path}")

        if not resources_dir:
            issues.append(f"Ruta de recursos vacía. Define {ENV_RESOURCES_DIR} o 'resources_dir' en config.")
        elif not os.path.isdir(resources_dir):
            issues.append(f"No existe el directorio de recursos en: {resources_dir}")

        return excel_path, resources_dir, info, issues

    # ============================================================
    # ✅ Orders (modo prueba local)
    # - Genera una clave única
    # - Guarda un JSON en "orders_outbox"
    # - Envía correo con el JSON adjunto
    # - Responde con order_key y número WhatsApp destino (SmartKet)
    # ============================================================

    def _generate_order_key() -> str:
        # Ej: SK-20260101-8F3KQ2 (token base32-ish)
        stamp = datetime.now(timezone.utc).strftime("%Y%m%d")
        token = secrets.token_urlsafe(6).replace("-", "").replace("_", "")[:6].upper()
        return f"SK-{stamp}-{token}"

    def _orders_outbox_dir(cfg: dict) -> Path:
        d = cfg.get("orders_outbox_dir") or os.path.join(os.path.dirname(__file__), "orders_outbox")
        p = Path(d).expanduser().resolve()
        p.mkdir(parents=True, exist_ok=True)
        return p

    def _send_order_email(cfg: dict, subject: str, body: str, attachment_name: str, attachment_bytes: bytes) -> None:
        smtp_host = cfg.get("smtp_host", "")
        smtp_port = int(cfg.get("smtp_port", 587))
        smtp_user = cfg.get("smtp_user", "")
        smtp_pass = cfg.get("smtp_pass", "")
        mail_to = cfg.get("mail_to", "")

        if not (smtp_host and smtp_user and smtp_pass and mail_to):
            raise ValueError("Config SMTP incompleta. Revisa smtp_host/smtp_user/smtp_pass/mail_to.")

        msg = EmailMessage()
        msg["Subject"] = subject
        msg["From"] = smtp_user
        msg["To"] = mail_to
        msg.set_content(body)

        msg.add_attachment(
            attachment_bytes,
            maintype="application",
            subtype="json",
            filename=attachment_name
        )

        with smtplib.SMTP(smtp_host, smtp_port) as s:
            s.ehlo()
            s.starttls()
            s.login(smtp_user, smtp_pass)
            s.send_message(msg)

    # ============================================================
    # ✅ CORS (Paso 1 - ajuste para que NO falle)
    # ============================================================

    @app.before_request
    def _handle_preflight():
        if request.method == "OPTIONS":
            # Respuesta vacía; headers los pone after_request.
            return make_response("", 204)

    @app.after_request
    def _add_cors_headers(resp):
        origin = request.headers.get("Origin")
        resp.headers["Access-Control-Allow-Origin"] = origin if origin else "*"
        resp.headers["Vary"] = "Origin"
        resp.headers["Access-Control-Allow-Methods"] = "GET, OPTIONS"
        resp.headers["Access-Control-Allow-Headers"] = "Content-Type"
        return resp

    @app.get("/api/health")
    def health():
        cfg = app.config["SMARTKET_CONFIG"]
        _excel, _res, info, issues = _get_effective_paths(cfg)

        return jsonify({
            "ok": True,
            "service": "smartket-publicweb-backend",
            "excel": {
                "path": info["excel_path"],
                "source": info["excel_source"],
                "exists": bool(info["excel_path"]) and os.path.isfile(info["excel_path"]),
            },
            "resources": {
                "dir": info["resources_dir"],
                "source": info["resources_source"],
                "exists": bool(info["resources_dir"]) and os.path.isdir(info["resources_dir"]),
            },
            "config_issues": issues,
        })

    @app.get("/api/recipes")
    def get_recipes():
        cfg = app.config["SMARTKET_CONFIG"]
        excel_path, resources_dir, info, issues = _get_effective_paths(cfg)

        if issues:
            return jsonify({
                "ok": False,
                "error": "config_invalid",
                "message": "Config inválida para leer recetas. Revisa rutas.",
                "details": issues,
                "excel_source": info["excel_source"],
                "resources_source": info["resources_source"],
            }), 500

        try:
            recipes = read_recipes_from_excel(excel_path)

            # ✅ Index recursivo de imágenes por ID (ignora subcarpetas)
            img_index = build_image_index(resources_dir)

            # ✅ Rellenar el campo img si existe imagen
            for r in recipes:
                rid = str(r.get("id", "")).strip()
                if rid and rid in img_index:
                    r["img"] = f"/api/images/{rid}"
                else:
                    r["img"] = None

            return jsonify({
                "ok": True,
                "count": len(recipes),
                "recipes": recipes
            })
        except Exception as ex:
            return jsonify({
                "ok": False,
                "error": str(ex)
            }), 500

    @app.get("/api/images/<recipe_id>")
    def get_image(recipe_id: str):
        cfg = app.config["SMARTKET_CONFIG"]
        _excel, resources_dir, info, issues = _get_effective_paths(cfg)

        # Para imágenes solo nos importa resources_dir
        img_issues = [
            x for x in issues
            if "recursos" in x.lower() or "resources" in x.lower() or "directorio" in x.lower()
        ]
        if img_issues:
            return jsonify({
                "ok": False,
                "error": "config_invalid",
                "message": "Config inválida para resolver imágenes. Revisa resources_dir.",
                "details": img_issues,
                "resources_source": info["resources_source"],
            }), 500

        img_index = build_image_index(resources_dir)
        rid = str(recipe_id).strip()

        if not rid or rid not in img_index:
            return jsonify({"ok": False, "error": "image_not_found", "id": rid}), 404

        path = img_index[rid]
        return send_file(path)

    

    @app.get("/api/ingredients/meta")
    def get_ingredients_meta():
        """Devuelve meta de Ingredientes.xlsx (por ahora: categoría) para UI."""
        cfg = app.config["SMARTKET_CONFIG"]
        excel_path, _res_dir, info, issues_paths = _get_effective_paths(cfg)
        if issues_paths:
            return jsonify({
                "ok": False,
                "error": "config_invalid",
                "details": issues_paths,
                "excel_source": info.get("excel_source"),
                "resources_source": info.get("resources_source"),
            }), 500

        ingredientes_xlsx = _derive_ingredientes_xlsx_path(excel_path)
        meta, issues_meta = _load_ingredients_meta(ingredientes_xlsx)

        return jsonify({
            "ok": True,
            "ingredients_xlsx": ingredientes_xlsx,
            "meta": meta,
            "issues": issues_meta,
        })


    @app.post("/api/orders/quote")
    def quote_order():
        """
        Cotiza el pedido en base al plan actual:
        - GRANEL=SI: prorratea por cantidad (precio de venta)
        - GRANEL=NO: vende empaques completos; elige empaque más chico para minimizar desperdicio
        """
        cfg = app.config["SMARTKET_CONFIG"]

        try:
            payload = request.get_json(silent=True) or {}
            excel_path, _res_dir, info, issues_paths = _get_effective_paths(cfg)
            if issues_paths:
                return jsonify({
                    "ok": False,
                    "error": "config_invalid",
                    "details": issues_paths,
                    "excel_source": info.get("excel_source"),
                    "resources_source": info.get("resources_source"),
                }), 500

            ingredientes_xlsx = _derive_ingredientes_xlsx_path(excel_path)
            catalog, issues_cat = _load_ingredients_catalog(ingredientes_xlsx)

            agg = _aggregate_plan_ingredients(payload)

            offer_overrides = payload.get("offerOverrides") or payload.get("offer_overrides") or {}
            items, total, issues_quote = _quote_sellable_items(agg, catalog, offer_overrides)

            purchase_snapshot = _build_purchase_snapshot_from_quote_items(items)

            return jsonify({
                "ok": True,
                "ingredients_xlsx": ingredientes_xlsx,
                "items": items,
                "order_total": total,
                "purchaseSnapshot": purchase_snapshot,
                "issues": issues_cat + issues_quote,
            })
        except Exception as ex:
            return jsonify({"ok": False, "error": str(ex)}), 500

    @app.post("/api/orders")
    def create_order():
        cfg = app.config["SMARTKET_CONFIG"]

        try:
            payload = request.get_json(silent=True) or {}
            order_key = _generate_order_key()


            # ------------------------------------------------------------------
            # Contrato JSON v2 (Iteración 1 + 2)
            # - recipeBlocks ya vienen POST-EXCLUSIÓN (con day + slot)
            # - se persisten offerOverrides del cliente (marca/presentación elegida)
            # - se incluye quoteSnapshot para que WPF pueda generar lista de compra
            # ------------------------------------------------------------------

            dynamic_days = payload.get("dynamicDays")
            excluded = payload.get("excludedIngredients", [])
            offer_overrides = payload.get("offerOverrides") or payload.get("offer_overrides") or {}

            days = _safe_days_from_dynamic_days(dynamic_days)
            recipe_blocks = _build_recipe_blocks(payload.get("plan"), dynamic_days, excluded)

            quote_snapshot = None
            quote_issues = []
            try:
                excel_path, _res_dir, info, issues_paths = _get_effective_paths(cfg)
                if issues_paths:
                    quote_issues.extend(issues_paths)
                else:
                    ingredientes_xlsx = _derive_ingredientes_xlsx_path(excel_path)
                    catalog, issues_cat = _load_ingredients_catalog(ingredientes_xlsx)
                    quote_issues.extend(issues_cat)

                    plan_for_quote = _build_plan_for_quote(payload.get("plan"), excluded)
                    agg = _aggregate_plan_ingredients({"plan": plan_for_quote})

                    items, total, issues_quote = _quote_sellable_items(agg, catalog, offer_overrides)

                    quote_issues.extend(issues_quote)

                    quote_snapshot = {
                        "ingredients_xlsx": ingredientes_xlsx,
                        "items": items,
                        "order_total": total,
                        "issues": quote_issues,
                    }
            except Exception as ex2:
                quote_issues.append(str(ex2))
                quote_snapshot = {
                    "items": [],
                    "order_total": 0.0,
                    "issues": quote_issues,
                }


            # purchaseSnapshot (contrato fuerte): se congela a partir de quoteSnapshot.items
            purchase_snapshot = _build_purchase_snapshot_from_quote_items((quote_snapshot or {}).get("items") or [])
            outbox = _orders_outbox_dir(cfg)

            order = {
                "order_key": order_key,
                "created_at_utc": datetime.now(timezone.utc).isoformat(),
                "deliveryType": payload.get("deliveryType"),
                # trazabilidad (raw)
                "plan": payload.get("plan"),
                "dynamicDays": payload.get("dynamicDays"),
                "excludedIngredients": excluded,
                "clientMeta": payload.get("clientMeta", {}),
                # contrato v2
                "days": days,
                "recipeBlocks": recipe_blocks,
                "offerOverrides": offer_overrides,
                "quoteSnapshot": quote_snapshot,
                "purchaseSnapshot": purchase_snapshot,
                "schema_version": 2
            }

            order_json = json.dumps(order, ensure_ascii=False, indent=2).encode("utf-8")
            file_path = outbox / f"{order_key}.json"
            file_path.write_bytes(order_json)

            # Email (modo prueba)
            subject = f"SmartKet Pedido {order_key}"
            body = (
                "Nuevo pedido SmartKet (modo pruebaa).\n\n"
                f"Clave: {order_key}\n"
                "Se adjunta archivo JSON para procesar en la app interna una vez confirmado el pago.\n"
            )
            _send_order_email(cfg, subject, body, f"{order_key}.json", order_json)

            # WhatsApp destino (SmartKet)
            wa = str(cfg.get("whatsapp_test_number", "")).strip() or "5562527059"

            return jsonify({
                "ok": True,
                "order_key": order_key,
                "whatsapp_number": wa
            })
        except Exception as ex:
            return jsonify({
                "ok": False,
                "error": str(ex)
            }), 500

    return app


def main():
    parser = argparse.ArgumentParser(description="SmartKet.PublicWeb local backend (Flask)")
    parser.add_argument("--config", required=True, help="Ruta a config.local.json")
    args = parser.parse_args()

    config = load_config(args.config)
    app = create_app(config)

    host = config.get("host", "127.0.0.1")
    port = int(config.get("port", 5050))

    app.run(host=host, port=port, debug=True)


if __name__ == "__main__":
    main()