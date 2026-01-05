#!/usr/bin/env python3
"""
SmartKet - Manual Excel -> Postgres sync (v1)

Corre esto manualmente cuando quieras “refrescar” Postgres con tus Excel.

Requiere:
  pip install openpyxl psycopg2-binary

Uso:
  python sync_to_postgres.py --recipes-dir "C:\\Users\\Proyectos\\Desktop\\Recetas"
  python sync_to_postgres.py --recipes-dir "C:\\Users\\Proyectos\\Desktop\\Recetas" --dry-run
  python sync_to_postgres.py --recipes-dir "C:\\Users\\Proyectos\\Desktop\\Recetas" --only recipes
  python sync_to_postgres.py --recipes-dir "C:\\Users\\Proyectos\\Desktop\\Recetas" --only ingredients
"""

from __future__ import annotations

import argparse
import os
import re
import sys
import unicodedata
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Dict, Iterable, List, Optional

try:
    import openpyxl  # type: ignore
except Exception:
    print("ERROR: Falta openpyxl. Instala con: pip install openpyxl", file=sys.stderr)
    raise

try:
    import psycopg2  # type: ignore
except Exception:
    print("ERROR: Falta psycopg2-binary. Instala con: pip install psycopg2-binary", file=sys.stderr)
    raise


_ws_re = re.compile(r"\s+")


def normalize_text(value: str) -> str:
    v = (value or "").strip().lower()
    v = _ws_re.sub(" ", v)
    v = "".join(ch for ch in unicodedata.normalize("NFD", v) if unicodedata.category(ch) != "Mn")
    return v


def split_pipe_list(value: Optional[object]) -> List[str]:
    if value is None:
        return []
    v = str(value).strip()
    if not v or normalize_text(v) == "na":
        return []
    return [item.strip() for item in v.split("|") if item.strip()]


def parse_decimal(value: Optional[object]) -> Optional[Decimal]:
    if value is None:
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value))
        except Exception:
            return None
    s = str(value).strip()
    if not s or normalize_text(s) == "na":
        return None
    s = s.replace(",", ".")
    try:
        return Decimal(s)
    except Exception:
        return None


def parse_int(value: Optional[object]) -> Optional[int]:
    d = parse_decimal(value)
    if d is None:
        return None
    try:
        return int(d)
    except Exception:
        return None


@dataclass(frozen=True)
class IngredientOffer:
    position: int
    brand: Optional[str]
    purchase_price: Optional[Decimal]
    sale_price: Optional[Decimal]
    presentation: Optional[str]
    unit: Optional[str]
    quantity: Optional[Decimal]


@dataclass(frozen=True)
class IngredientRow:
    name: str
    category: Optional[str]
    offers: List[IngredientOffer]
    forms: List[str]
    processes: List[str]


@dataclass(frozen=True)
class RecipeIngredientLine:
    ingredient_name_raw: str
    unit_raw: Optional[str]
    quantity: Optional[Decimal]
    form_raw: Optional[str]
    process_raw: Optional[str]
    position: int


@dataclass(frozen=True)
class RecipeBlock:
    recipe_id: str
    title: str
    calories: Optional[int]
    prep_time_minutes: Optional[int]
    servings: Optional[Decimal]
    preparation_text: Optional[str]
    meal_bucket: Optional[str]
    source_excel_sheet: Optional[str]
    source_row_hint: Optional[int]
    ingredients: List[RecipeIngredientLine]


def read_ingredientes_xlsx(path: Path) -> List[IngredientRow]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    rows: List[IngredientRow] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        ingredient = (row[0] or "").strip() if row[0] else ""
        if not ingredient:
            continue

        marcas = split_pipe_list(row[1])
        compras = split_pipe_list(row[2])
        ventas = split_pipe_list(row[3])
        presentaciones = split_pipe_list(row[4])
        unidades = split_pipe_list(row[5])
        cantidades = split_pipe_list(row[6])

        category = str(row[7]).strip() if row[7] is not None and str(row[7]).strip() else None
        forms = split_pipe_list(row[8])
        processes = split_pipe_list(row[9])

        max_len = max(len(marcas), len(compras), len(ventas), len(presentaciones), len(unidades), len(cantidades), 0)

        offers: List[IngredientOffer] = []
        for pos in range(max_len):
            brand = marcas[pos] if pos < len(marcas) else None
            purchase = parse_decimal(compras[pos]) if pos < len(compras) else None
            sale = parse_decimal(ventas[pos]) if pos < len(ventas) else None
            presentation = presentaciones[pos] if pos < len(presentaciones) else None
            unit = unidades[pos] if pos < len(unidades) else None
            qty = parse_decimal(cantidades[pos]) if pos < len(cantidades) else None

            if not any([brand, purchase, sale, presentation, unit, qty]):
                continue

            offers.append(
                IngredientOffer(
                    position=pos,
                    brand=(brand if brand and normalize_text(brand) != "na" else None),
                    purchase_price=purchase,
                    sale_price=sale,
                    presentation=presentation,
                    unit=unit,
                    quantity=qty,
                )
            )

        rows.append(IngredientRow(name=ingredient, category=category, offers=offers, forms=forms, processes=processes))

    return rows


def read_alimentos_xlsx(path: Path) -> List[RecipeBlock]:
    wb = openpyxl.load_workbook(path, data_only=True)
    recipes: List[RecipeBlock] = []

    for ws in wb.worksheets:
        sheet_name = ws.title
        meal_bucket = normalize_text(sheet_name)

        r = 1
        max_r = ws.max_row or 1

        while r <= max_r:
            a = ws.cell(row=r, column=1).value
            if (normalize_text(str(a)) if a is not None else "") != "receta":
                r += 1
                continue

            title = str(ws.cell(row=r, column=2).value or "").strip()
            recipe_id = str(ws.cell(row=r, column=3).value or "").strip()
            calories = parse_int(ws.cell(row=r, column=4).value)
            prep_mins = parse_int(ws.cell(row=r, column=5).value)
            servings = parse_decimal(ws.cell(row=r, column=6).value)

            prep_text = None
            r_prep = r + 1
            if r_prep <= max_r and normalize_text(str(ws.cell(row=r_prep, column=1).value or "")) == "preparacion":
                prep_text = ws.cell(row=r_prep, column=2).value
                prep_text = str(prep_text).strip() if prep_text is not None else None

            r_ing_header = r + 2
            if r_ing_header <= max_r and normalize_text(str(ws.cell(row=r_ing_header, column=1).value or "")) != "ingrediente":
                found = False
                for k in range(r + 1, min(r + 7, max_r + 1)):
                    if normalize_text(str(ws.cell(row=k, column=1).value or "")) == "ingrediente":
                        r_ing_header = k
                        found = True
                        break
                if not found:
                    r += 1
                    continue

            ingredient_lines: List[RecipeIngredientLine] = []
            rr = r_ing_header + 1
            pos = 0
            while rr <= max_r:
                ing_name = ws.cell(row=rr, column=1).value
                if ing_name is None or str(ing_name).strip() == "":
                    break

                ing_name_raw = str(ing_name).strip()
                unit_raw = ws.cell(row=rr, column=2).value
                unit_raw = str(unit_raw).strip() if unit_raw is not None and str(unit_raw).strip() else None
                qty = parse_decimal(ws.cell(row=rr, column=3).value)
                form_raw = ws.cell(row=rr, column=4).value
                form_raw = str(form_raw).strip() if form_raw is not None and str(form_raw).strip() else None
                process_raw = ws.cell(row=rr, column=5).value
                process_raw = str(process_raw).strip() if process_raw is not None and str(process_raw).strip() else None

                ingredient_lines.append(
                    RecipeIngredientLine(
                        ingredient_name_raw=ing_name_raw,
                        unit_raw=unit_raw,
                        quantity=qty,
                        form_raw=form_raw,
                        process_raw=process_raw,
                        position=pos,
                    )
                )
                pos += 1
                rr += 1

            if recipe_id and title:
                recipes.append(
                    RecipeBlock(
                        recipe_id=recipe_id,
                        title=title,
                        calories=calories,
                        prep_time_minutes=prep_mins,
                        servings=servings,
                        preparation_text=prep_text,
                        meal_bucket=meal_bucket,
                        source_excel_sheet=sheet_name,
                        source_row_hint=r,
                        ingredients=ingredient_lines,
                    )
                )

            r = rr + 1

    return recipes


IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp"}


def index_recipe_images(resources_dir: Path) -> Dict[str, str]:
    if not resources_dir.exists():
        return {}
    matches: Dict[str, str] = {}
    for p in sorted(resources_dir.rglob("*")):
        if not p.is_file():
            continue
        if p.suffix.lower() not in IMAGE_EXTS:
            continue
        stem = p.stem.strip()
        if stem.isdigit() and stem not in matches:
            matches[stem] = p.name
    return matches


def pg_connect():
    host = os.getenv("PGHOST", "localhost")
    port = int(os.getenv("PGPORT", "5432"))
    db = os.getenv("PGDATABASE", "smartket")
    user = os.getenv("PGUSER", "smartket")
    pwd = os.getenv("PGPASSWORD", "smartket")
    return psycopg2.connect(host=host, port=port, dbname=db, user=user, password=pwd)


def upsert_ingredient(cur, row: IngredientRow) -> int:
    name_norm = normalize_text(row.name)
    cur.execute(
        """
        INSERT INTO ingredients (name, name_normalized, category)
        VALUES (%s, %s, %s)
        ON CONFLICT (name_normalized)
        DO UPDATE SET name = EXCLUDED.name, category = EXCLUDED.category
        RETURNING id;
        """,
        (row.name, name_norm, row.category),
    )
    return int(cur.fetchone()[0])


def ensure_form(cur, name: str) -> int:
    n = name.strip()
    nn = normalize_text(n)
    cur.execute(
        """
        INSERT INTO forms (name, name_normalized)
        VALUES (%s, %s)
        ON CONFLICT (name_normalized) DO UPDATE SET name = EXCLUDED.name
        RETURNING id;
        """,
        (n, nn),
    )
    return int(cur.fetchone()[0])


def ensure_process(cur, name: str) -> int:
    n = name.strip()
    nn = normalize_text(n)
    cur.execute(
        """
        INSERT INTO processes (name, name_normalized)
        VALUES (%s, %s)
        ON CONFLICT (name_normalized) DO UPDATE SET name = EXCLUDED.name
        RETURNING id;
        """,
        (n, nn),
    )
    return int(cur.fetchone()[0])


def reconcile_links(cur, table: str, left_col: str, right_col: str, left_id: int, desired_right_ids: Iterable[int]):
    desired = set(int(x) for x in desired_right_ids)
    cur.execute(f"SELECT {right_col} FROM {table} WHERE {left_col}=%s;", (left_id,))
    existing = set(int(r[0]) for r in cur.fetchall())

    for rid in existing - desired:
        cur.execute(f"DELETE FROM {table} WHERE {left_col}=%s AND {right_col}=%s;", (left_id, rid))
    for rid in desired - existing:
        cur.execute(
            f"INSERT INTO {table} ({left_col}, {right_col}) VALUES (%s,%s) ON CONFLICT DO NOTHING;",
            (left_id, rid),
        )


def upsert_offer(cur, ingredient_id: int, offer: IngredientOffer):
    cur.execute(
        """
        INSERT INTO ingredient_offers
          (ingredient_id, position, brand, purchase_price, sale_price, presentation, unit, quantity)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
        ON CONFLICT (ingredient_id, position)
        DO UPDATE SET
          brand = EXCLUDED.brand,
          purchase_price = EXCLUDED.purchase_price,
          sale_price = EXCLUDED.sale_price,
          presentation = EXCLUDED.presentation,
          unit = EXCLUDED.unit,
          quantity = EXCLUDED.quantity;
        """,
        (
            ingredient_id,
            offer.position,
            offer.brand,
            offer.purchase_price,
            offer.sale_price,
            offer.presentation,
            offer.unit,
            offer.quantity,
        ),
    )


def delete_offers_after_position(cur, ingredient_id: int, max_position_exclusive: int):
    cur.execute(
        "DELETE FROM ingredient_offers WHERE ingredient_id=%s AND position >= %s;",
        (ingredient_id, max_position_exclusive),
    )


def upsert_recipe(cur, r: RecipeBlock):
    cur.execute(
        """
        INSERT INTO recipes
          (id, title, calories, prep_time_minutes, servings, preparation_text, meal_bucket, source_excel_sheet, source_row_hint)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
        ON CONFLICT (id)
        DO UPDATE SET
          title = EXCLUDED.title,
          calories = EXCLUDED.calories,
          prep_time_minutes = EXCLUDED.prep_time_minutes,
          servings = EXCLUDED.servings,
          preparation_text = EXCLUDED.preparation_text,
          meal_bucket = EXCLUDED.meal_bucket,
          source_excel_sheet = EXCLUDED.source_excel_sheet,
          source_row_hint = EXCLUDED.source_row_hint;
        """,
        (
            r.recipe_id,
            r.title,
            r.calories,
            r.prep_time_minutes,
            r.servings,
            r.preparation_text,
            r.meal_bucket,
            r.source_excel_sheet,
            r.source_row_hint,
        ),
    )


def ingredient_id_by_name(cur, ingredient_name_raw: str) -> Optional[int]:
    nn = normalize_text(ingredient_name_raw)
    cur.execute("SELECT id FROM ingredients WHERE name_normalized=%s;", (nn,))
    row = cur.fetchone()
    return int(row[0]) if row else None


def replace_recipe_ingredients(cur, recipe_id: str, lines: List[RecipeIngredientLine]):
    cur.execute("DELETE FROM recipe_ingredients WHERE recipe_id=%s;", (recipe_id,))
    for line in lines:
        ing_id = ingredient_id_by_name(cur, line.ingredient_name_raw)
        cur.execute(
            """
            INSERT INTO recipe_ingredients
              (recipe_id, ingredient_name_raw, ingredient_id, unit_raw, quantity, form_raw, process_raw, position)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s);
            """,
            (
                recipe_id,
                line.ingredient_name_raw,
                ing_id,
                line.unit_raw,
                line.quantity,
                line.form_raw,
                line.process_raw,
                line.position,
            ),
        )


def upsert_recipe_media_primary(cur, recipe_id: str, filename: str):
    cur.execute("DELETE FROM recipe_media WHERE recipe_id=%s AND is_primary=TRUE;", (recipe_id,))
    cur.execute(
        """
        INSERT INTO recipe_media (recipe_id, media_type, filename, is_primary, source)
        VALUES (%s,'image',%s,TRUE,'excel_import');
        """,
        (recipe_id, filename),
    )


def run_sync(recipes_dir: Path, *, only: str = "all", dry_run: bool = False) -> None:
    alimentos_path = recipes_dir / "Alimentos.xlsx"
    ingredientes_path = recipes_dir / "Ingredientes.xlsx"
    recursos_dir = recipes_dir / "Recursos"

    if not alimentos_path.exists():
        raise FileNotFoundError(f"Missing {alimentos_path}")
    if not ingredientes_path.exists():
        raise FileNotFoundError(f"Missing {ingredientes_path}")

    print(f"[1/4] Reading Ingredientes.xlsx: {ingredientes_path}")
    ingredient_rows = read_ingredientes_xlsx(ingredientes_path)
    print(f"      -> {len(ingredient_rows)} ingredient rows")

    print(f"[2/4] Reading Alimentos.xlsx: {alimentos_path}")
    recipe_blocks = read_alimentos_xlsx(alimentos_path)
    print(f"      -> {len(recipe_blocks)} recipes found")

    print(f"[3/4] Indexing images in Recursos/: {recursos_dir}")
    image_index = index_recipe_images(recursos_dir)
    print(f"      -> {len(image_index)} recipe images matched by ID")

    if dry_run:
        print("[4/4] DRY-RUN: No database changes will be made.")
        print("      Would sync sections:", only)
        print("      - ingredients rows:", len(ingredient_rows))
        print("      - recipes found:", len(recipe_blocks))
        print("      - recipe images matched:", len(image_index))
        print("\n✅ DRY-RUN DONE. Nothing was written to Postgres.")
        return

    print("[4/4] Syncing into PostgreSQL (single transaction)...")
    with pg_connect() as conn:
        conn.autocommit = False
        with conn.cursor() as cur:
            do_ingredients = (only in ("all", "ingredients"))
            do_recipes = (only in ("all", "recipes"))

            if not (do_ingredients or do_recipes):
                raise ValueError(f"Invalid --only value: {only!r}")
            if do_ingredients:
                for ing in ingredient_rows:
                    ing_id = upsert_ingredient(cur, ing)

                    for offer in ing.offers:
                        upsert_offer(cur, ing_id, offer)
                    delete_offers_after_position(cur, ing_id, max_position_exclusive=len(ing.offers))

                    form_ids = [ensure_form(cur, f) for f in ing.forms]
                    process_ids = [ensure_process(cur, p) for p in ing.processes]
                    reconcile_links(cur, "ingredient_forms", "ingredient_id", "form_id", ing_id, form_ids)
                    reconcile_links(cur, "ingredient_processes", "ingredient_id", "process_id", ing_id, process_ids)

            if do_recipes:
                for r in recipe_blocks:
                    upsert_recipe(cur, r)
                    replace_recipe_ingredients(cur, r.recipe_id, r.ingredients)

                    filename = image_index.get(r.recipe_id)
                    if filename:
                        upsert_recipe_media_primary(cur, r.recipe_id, filename)

        conn.commit()

    print("\n✅ DONE. Postgres is now synced to the current Excel state.")
    print("   (Manual sync: it won't update unless you run this script again.)")


def main() -> int:
    parser = argparse.ArgumentParser(description="SmartKet manual Excel->Postgres sync")
    parser.add_argument("--recipes-dir", required=False, help="Folder containing Alimentos.xlsx, Ingredientes.xlsx and Recursos/")
    parser.add_argument("--only", choices=["all", "ingredients", "recipes"], default="all", help="Which section to sync into Postgres")
    parser.add_argument("--dry-run", action="store_true", help="Parse Excel and show what would be synced, but do not write to Postgres")
    args = parser.parse_args()

    recipes_dir = args.recipes_dir or os.getenv("SMARTKET_RECIPES_DIR")
    if not recipes_dir:
        print("ERROR: Provide --recipes-dir or set SMARTKET_RECIPES_DIR.", file=sys.stderr)
        return 2

    run_sync(Path(recipes_dir).expanduser().resolve(), only=args.only, dry_run=args.dry_run)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())